"""
CNA Engine — Reference Data
Static game data loaded from CNA_DATA_TABLES.xlsx.
These are lookup tables used by the engine — they never change during play.
"""
from __future__ import annotations
from dataclasses import dataclass, field
from typing import Optional
import json
import os


# ════════════════════════════════════════
# TERRAIN EFFECTS
# ════════════════════════════════════════

@dataclass
class TerrainInfo:
    terrain_type: str
    motorized_cp: float
    non_motorized_cp: float
    track_effect: str       # "Halve" or "N/A"
    barrage_shift: str      # "—", "1L", "2L"
    anti_armor_shift: str
    close_assault_shift: str
    stacking_limit: str     # "Unlimited" or "5 SP" etc.
    notes: str = ""

    @property
    def barrage_shift_val(self) -> int:
        if self.barrage_shift in ("—", "-", ""):
            return 0
        return -int(self.barrage_shift[0])  # "1L" -> -1, "2L" -> -2

    @property
    def anti_armor_shift_val(self) -> int:
        if self.anti_armor_shift in ("—", "-", ""):
            return 0
        return -int(self.anti_armor_shift[0])

    @property
    def close_assault_shift_val(self) -> int:
        if self.close_assault_shift in ("—", "-", ""):
            return 0
        return -int(self.close_assault_shift[0])

    @property
    def track_cp_motorized(self) -> Optional[int]:
        if self.track_effect == "Halve":
            return max(1, self.motorized_cp // 2)
        return None  # Tracks don't apply


# ════════════════════════════════════════
# CPA COSTS
# ════════════════════════════════════════

@dataclass
class CPACost:
    action: str
    cp_cost: str            # Could be int or "Varies"
    notes: str = ""


# ════════════════════════════════════════
# PORTS
# ════════════════════════════════════════

@dataclass
class PortInfo:
    name: str
    max_efficiency: str     # int or "∞"
    max_tonnage_per_gt: str # int or "∞"
    sp_capacity: str
    repair_facility: str
    notes: str = ""


# ════════════════════════════════════════
# AIRCRAFT CHARACTERISTICS
# ════════════════════════════════════════

@dataclass
class AircraftCharacteristics:
    name: str
    side: str
    njh_type: str
    maneuver: str           # Could be "30/32" for Bf.110
    tacair: str
    bombload: str
    range_hexes: int
    fuel_consumption: int
    maintenance: str
    d_capable: bool
    night_capable: bool
    transport_capacity: str
    errata_notes: str = ""


# ════════════════════════════════════════
# INITIATIVE RATINGS
# ════════════════════════════════════════

@dataclass
class InitiativeRating:
    gt_range: str
    gt_start: int
    gt_end: int
    allied_rating: int
    axis_rating: int


# ════════════════════════════════════════
# EVAPORATION RATES
# ════════════════════════════════════════

@dataclass
class EvaporationRate:
    period: str
    gt_threshold: int       # Rate changes at this GT
    allied_rate: float      # 0.09 = 9%
    axis_rate: float


# ════════════════════════════════════════
# KEY DATES / EVENTS
# ════════════════════════════════════════

@dataclass
class KeyDate:
    game_turn: str
    date: str
    event: str
    rule_ref: str = ""


# ════════════════════════════════════════
# REFERENCE DATA CONTAINER
# ════════════════════════════════════════

@dataclass
class ReferenceData:
    """All static reference data the engine needs."""
    terrain: dict[str, TerrainInfo] = field(default_factory=dict)
    cpa_costs: list[CPACost] = field(default_factory=list)
    ports: dict[str, PortInfo] = field(default_factory=dict)
    aircraft: dict[str, AircraftCharacteristics] = field(default_factory=dict)
    initiative_ratings: list[InitiativeRating] = field(default_factory=list)
    evaporation_rates: list[EvaporationRate] = field(default_factory=list)
    key_dates: list[KeyDate] = field(default_factory=list)

    # CRT tables stored as nested dicts for lookup
    barrage_crt: dict = field(default_factory=dict)
    anti_armor_crt: dict = field(default_factory=dict)
    close_assault_crt: dict = field(default_factory=dict)
    flak_crt: dict = field(default_factory=dict)
    air_bombardment_table: dict = field(default_factory=dict)
    strafing_table: dict = field(default_factory=dict)
    morale_table: dict = field(default_factory=dict)
    fuel_consumption_chart: dict = field(default_factory=dict)
    ammo_consumption: dict = field(default_factory=dict)

    def get_initiative_rating(self, game_turn: int):
        """Get initiative ratings for a specific game turn."""
        for ir in self.initiative_ratings:
            if ir.gt_start <= game_turn <= ir.gt_end:
                return ir.allied_rating, ir.axis_rating
        # Default to last known
        return 5, 4

    def get_evaporation_rates(self, game_turn: int):
        """Get evaporation rates for a specific game turn."""
        allied_rate = 0.09  # Default pre-jerrycan
        axis_rate = 0.03
        for er in self.evaporation_rates:
            if game_turn >= er.gt_threshold:
                allied_rate = er.allied_rate
                axis_rate = er.axis_rate
        return allied_rate, axis_rate

    def get_terrain_movement_cost(self, terrain_type: str, is_motorized: bool,
                                   has_track: bool = False) -> float:
        """Get movement CP cost for terrain."""
        info = self.terrain.get(terrain_type)
        if not info:
            return 2  # Default to clear
        base = info.motorized_cp if is_motorized else info.non_motorized_cp
        if has_track and info.track_effect == "Halve":
            return max(1, base / 2)
        return base

    def get_stacking_limit(self, terrain_type: str) -> Optional[int]:
        """Get stacking limit in SP for a terrain type. None = unlimited."""
        info = self.terrain.get(terrain_type)
        if not info:
            return None  # Unknown terrain → no limit
        s = info.stacking_limit.strip()
        if s.lower() == "unlimited" or not s:
            return None
        # Parse "10 SP", "5 SP", etc.
        parts = s.split()
        if parts and parts[0].isdigit():
            return int(parts[0])
        try:
            return int(s)
        except (ValueError, TypeError):
            return None

    def lookup_fuel_consumption(self, cps_expended: int, fuel_rate: int) -> float:
        """Look up fuel consumed from the fuel consumption chart."""
        # Formula: fuel = cps_expended * fuel_rate * 0.2
        return round(cps_expended * fuel_rate * 0.2, 1)


def load_reference_data_from_xlsx(filepath: str) -> ReferenceData:
    """Load all reference data from the CNA_DATA_TABLES.xlsx file."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ref = ReferenceData()

    # ── Terrain Effects ──
    ws = wb["Terrain Effects"]
    skip_labels = {"Terrain Type", "ERRATA NOTES:", "HEXSIDE EFFECTS", "Hexside Type",
                   "Yellow rows = errata-corrected values"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0] or str(row[0]) in skip_labels or str(row[0]).startswith("["):
            continue
        key = str(row[0]).lower().replace(" ", "_")
        # Handle non-integer CPs (Road=0.5, Railroad="Rail only")
        def _parse_cp(val):
            if val is None:
                return 0
            if isinstance(val, (int, float)):
                return val
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0  # "Rail only" etc.
        ref.terrain[key] = TerrainInfo(
            terrain_type=str(row[0]),
            motorized_cp=_parse_cp(row[1]),
            non_motorized_cp=_parse_cp(row[2]),
            track_effect=str(row[3] or "N/A"),
            barrage_shift=str(row[4] or "—"),
            anti_armor_shift=str(row[5] or "—"),
            close_assault_shift=str(row[6] or "—"),
            stacking_limit=str(row[7] or "Unlimited"),
            notes=str(row[8] or ""),
        )

    # ── CPA Costs ──
    ws = wb["CPA Costs"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0]:
            ref.cpa_costs.append(CPACost(
                action=str(row[0]),
                cp_cost=str(row[1] or ""),
                notes=str(row[2] or ""),
            ))

    # ── Ports ──
    ws = wb["Ports"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0] and not str(row[0]).startswith("["):
            ref.ports[str(row[0]).lower()] = PortInfo(
                name=str(row[0]),
                max_efficiency=str(row[1] or ""),
                max_tonnage_per_gt=str(row[2] or ""),
                sp_capacity=str(row[3] or ""),
                repair_facility=str(row[4] or ""),
                notes=str(row[5] or ""),
            )

    # ── Aircraft ──
    ws = wb["Aircraft"]
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row[0]:
            key = str(row[0]).lower().replace(" ", "_").replace(".", "")
            ref.aircraft[key] = AircraftCharacteristics(
                name=str(row[0]),
                side=str(row[1] or ""),
                njh_type=str(row[2] or ""),
                maneuver=str(row[3] or ""),
                tacair=str(row[4] or ""),
                bombload=str(row[5] or ""),
                range_hexes=int(row[6]) if row[6] and str(row[6]).isdigit() else 0,
                fuel_consumption=int(row[7]) if row[7] and str(row[7]).isdigit() else 0,
                maintenance=str(row[8] or ""),
                d_capable=str(row[9] or "").lower() == "yes",
                night_capable=str(row[10] or "").lower() == "yes",
                transport_capacity=str(row[11] or ""),
                errata_notes=str(row[12] or ""),
            )

    # ── Initiative ──
    ws = wb["Initiative"]
    for row in ws.iter_rows(min_row=4, max_row=7, values_only=True):
        if row[0] and row[1]:
            gt_text = str(row[0])
            # Parse GT range from text like "GT 1–42 (Sept 1940 – June 1941)"
            import re
            match = re.search(r'GT\s*(\d+)[–-](\d+)', gt_text)
            if match:
                ref.initiative_ratings.append(InitiativeRating(
                    gt_range=gt_text,
                    gt_start=int(match.group(1)),
                    gt_end=int(match.group(2)),
                    allied_rating=int(row[1]),
                    axis_rating=int(row[2]),
                ))

    # ── Evaporation ──
    ref.evaporation_rates = [
        EvaporationRate(period="Pre-jerrycan", gt_threshold=1, allied_rate=0.09, axis_rate=0.03),
        EvaporationRate(period="Post-jerrycan", gt_threshold=47, allied_rate=0.06, axis_rate=0.03),
    ]

    # ── Key Dates ──
    ws = wb["Key Dates"]
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0]:
            ref.key_dates.append(KeyDate(
                game_turn=str(row[0]),
                date=str(row[1] or ""),
                event=str(row[2] or ""),
                rule_ref=str(row[3] or ""),
            ))

    # ── Fuel Consumption Chart ──
    if "Logistics Tables" in wb.sheetnames:
        ws = wb["Logistics Tables"]
        ref.fuel_consumption_chart = {}
        for row in ws.iter_rows(min_row=6, max_row=20, values_only=True):
            if row[0] and isinstance(row[0], (int, float)):
                cp = int(row[0])
                ref.fuel_consumption_chart[cp] = {}
                for i, rate in enumerate(range(1, 8)):
                    if row[i + 1] is not None:
                        ref.fuel_consumption_chart[cp][rate] = float(row[i + 1])

    # ── Ammo Consumption ──
    ref.ammo_consumption = {
        "barrage": 4,
        "anti_armor": 3,
        "close_assault": 2,
        "anti_air": 2,
        "rearm_tacair": 1,
        "rearm_bombs": 1,  # per bombload
    }

    wb.close()
    return ref


def save_reference_data_json(ref: ReferenceData, filepath: str):
    """Save reference data as JSON for quick loading without xlsx dependency."""
    from dataclasses import fields, is_dataclass

    def _dc_to_dict(obj):
        if is_dataclass(obj) and not isinstance(obj, type):
            result = {}
            for f in fields(obj):
                result[f.name] = _dc_to_dict(getattr(obj, f.name))
            return result
        elif isinstance(obj, dict):
            return {k: _dc_to_dict(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [_dc_to_dict(item) for item in obj]
        elif isinstance(obj, (str, int, float, bool, type(None))):
            return obj
        return str(obj)

    data = _dc_to_dict(ref)
    os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)


def load_reference_data_json(filepath: str) -> dict:
    """Load reference data from cached JSON."""
    with open(filepath, "r") as f:
        return json.load(f)
