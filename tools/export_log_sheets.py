#!/usr/bin/env python3
"""
export_log_sheets.py  --  CNA Game Report Exporter

Reads game save files (saves/gt{N}.json) and game logs (logs/game_*.jsonl)
and produces an Excel workbook matching the CNA reference-sheet style.

Sheets produced:
  1. Unit Status   -- one row per unit per game turn (from save files)
  2. Orders        -- one row per order issued (from game log)
  3. Turn Summary  -- one row per turn with aggregate stats
  4. Dumps         -- supply dump snapshot from latest save

Usage:
  python3 tools/export_log_sheets.py [--log-dir logs] [--save-dir saves] [--output output/game_report.xlsx]
"""

import argparse
import glob
import json
import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl is required.  Install with:  pip install openpyxl")


# ---------------------------------------------------------------------------
#  Colour palette / styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL_ALLIED = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
HEADER_FILL_AXIS = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
HEADER_FILL_NEUTRAL = PatternFill(start_color="3A3A3A", end_color="3A3A3A", fill_type="solid")
ROW_FILL_ALLIED = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ROW_FILL_AXIS = PatternFill(start_color="F2D4D4", end_color="F2D4D4", fill_type="solid")
ROW_FILL_DESTROYED = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
PCT_FMT = "0.0%"
NUM1_FMT = "0.0"


# ---------------------------------------------------------------------------
#  Helpers
# ---------------------------------------------------------------------------
def _style_header(ws, row, fill, ncols):
    """Apply header styling to a row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(cell, is_number=False):
    cell.border = THIN_BORDER
    if is_number:
        cell.alignment = Alignment(horizontal="right")
    else:
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, min_width=8, max_width=30):
    """Auto-fit column widths (approximate)."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _sum_strength(strength_dict):
    """Sum a strength dict {infantry:N, armor:N, ...} -> total SP."""
    if not strength_dict or not isinstance(strength_dict, dict):
        return 0
    return sum(v for v in strength_dict.values() if isinstance(v, (int, float)))


def _safe_get(d, *keys, default=None):
    """Nested dict get."""
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur


# ---------------------------------------------------------------------------
#  Data loading
# ---------------------------------------------------------------------------
def discover_save_files(save_dir):
    """Return dict {turn_number: filepath} for every gt*.json (excluding _memory)."""
    pattern = os.path.join(save_dir, "gt*.json")
    files = glob.glob(pattern)
    result = {}
    for fp in files:
        basename = os.path.basename(fp)
        if "_memory" in basename:
            continue
        m = re.match(r"gt(\d+)\.json$", basename)
        if m:
            result[int(m.group(1))] = fp
    return result


def load_save(filepath):
    """Load a save JSON, return dict or None on failure."""
    try:
        with open(filepath, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"  WARNING: could not load {filepath}: {e}")
        return None


def discover_game_log(log_dir):
    """Return the path to the latest game_*.jsonl file, or None."""
    pattern = os.path.join(log_dir, "game_*.jsonl")
    files = sorted(glob.glob(pattern))
    return files[-1] if files else None


def load_game_log(filepath):
    """Load all entries from a JSONL game log. Return list of dicts."""
    entries = []
    try:
        with open(filepath, "r") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entries.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    except Exception as e:
        print(f"  WARNING: could not load game log {filepath}: {e}")
    return entries


# ---------------------------------------------------------------------------
#  Sheet 1: Unit Status
# ---------------------------------------------------------------------------
UNIT_STATUS_HEADERS = [
    "GT", "Side", "Nationality", "Unit Name", "Unit ID", "Class", "Size",
    "Hex", "Status",
    "SP", "TOE SP", "SP %",
    "Water", "Water Cap", "Water %",
    "Fuel", "Fuel Cap", "Fuel %",
    "Ammo", "Ammo Cap",
    "Stores", "Stores Cap",
    "CPA Base", "CPA Used", "Cohesion",
    "In Contact", "Trucks",
]


def write_unit_status(wb, saves_by_turn):
    ws = wb.create_sheet("Unit Status")

    # Header
    for ci, h in enumerate(UNIT_STATUS_HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, HEADER_FILL_NEUTRAL, len(UNIT_STATUS_HEADERS))

    row = 2
    for gt in sorted(saves_by_turn.keys()):
        data = load_save(saves_by_turn[gt])
        if data is None:
            continue
        turn_num = _safe_get(data, "turn", "game_turn", default=gt)
        units = data.get("units", {})

        # Sort: allied first, then axis; within side sort by name
        sorted_units = sorted(
            units.values(),
            key=lambda u: (0 if u.get("side") == "allied" else 1, u.get("name", "")),
        )

        for u in sorted_units:
            sp = _sum_strength(u.get("current_strength"))
            toe_sp = _sum_strength(u.get("toe_strength"))
            sp_pct = sp / toe_sp if toe_sp else 0

            supply = u.get("supply", {})
            water = supply.get("water", 0) or 0
            water_cap = supply.get("water_capacity", 0) or 0
            fuel = supply.get("fuel", 0) or 0
            fuel_cap = supply.get("fuel_capacity", 0) or 0
            ammo = supply.get("ammo", 0) or 0
            ammo_cap = supply.get("ammo_capacity", 0) or 0
            stores = supply.get("stores", 0) or 0
            stores_cap = supply.get("stores_capacity", 0) or 0

            water_pct = water / water_cap if water_cap else 0
            fuel_pct = fuel / fuel_cap if fuel_cap else 0

            values = [
                turn_num,
                u.get("side", ""),
                u.get("nationality", ""),
                u.get("name", ""),
                u.get("id", ""),
                u.get("unit_class", ""),
                u.get("unit_size", ""),
                u.get("hex_id", ""),
                u.get("status", ""),
                sp,
                toe_sp,
                sp_pct,
                round(water, 2),
                water_cap,
                water_pct,
                round(fuel, 2),
                fuel_cap,
                fuel_pct,
                round(ammo, 2),
                ammo_cap,
                round(stores, 2),
                stores_cap,
                u.get("base_cpa", 0),
                u.get("current_cpa_spent", 0),
                u.get("cohesion", 0),
                "Y" if u.get("is_in_contact") else "N",
                u.get("attached_truck_points", 0),
            ]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                _style_data_cell(cell, is_number=isinstance(v, (int, float)))

            # Row colouring
            status = u.get("status", "active")
            side = u.get("side", "")
            if status in ("destroyed", "withdrawn", "surrendered"):
                fill = ROW_FILL_DESTROYED
            elif side == "allied":
                fill = ROW_FILL_ALLIED
            else:
                fill = ROW_FILL_AXIS
            for ci in range(1, len(values) + 1):
                ws.cell(row=row, column=ci).fill = fill

            # Percentage formatting
            for pct_col in [12, 15, 18]:  # SP%, Water%, Fuel%
                ws.cell(row=row, column=pct_col).number_format = PCT_FMT

            row += 1

    # Freeze header
    ws.freeze_panes = "A2"
    _auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(UNIT_STATUS_HEADERS))}{row - 1}"
    return row - 2  # data rows written


# ---------------------------------------------------------------------------
#  Sheet 2: Orders
# ---------------------------------------------------------------------------
ORDERS_HEADERS = [
    "GT", "OpStage", "Phase", "SubPhase", "Side",
    "Command", "Unit ID", "Target/Destination", "Success", "Error",
]


def write_orders(wb, log_entries):
    ws = wb.create_sheet("Orders")

    for ci, h in enumerate(ORDERS_HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, HEADER_FILL_NEUTRAL, len(ORDERS_HEADERS))

    row = 2
    for entry in log_entries:
        if entry.get("type") != "phase":
            continue
        gt = entry.get("game_turn", "")
        op_stage = entry.get("op_stage", "")
        phase = entry.get("phase", "")
        sub_phase = entry.get("sub_phase", "")
        side = entry.get("side", "")

        orders = entry.get("orders", [])
        if not orders:
            continue

        for order in orders:
            cmd = order.get("command", "")
            params = order.get("params", {})

            # Extract unit_id from params (various keys)
            unit_id = (
                params.get("unit_id")
                or params.get("aircraft_id")
                or params.get("formation_id")
                or ""
            )

            # Extract target / destination
            target = (
                params.get("destination")
                or params.get("target_hex")
                or params.get("dump_id")
                or params.get("location")
                or params.get("target_unit_id")
                or params.get("mission")
                or ""
            )
            # If there are still interesting params, append them
            extra_keys = {"supplies", "fuel", "water", "ammo", "stores"}
            extras = {k: v for k, v in params.items()
                      if k in extra_keys and v}
            if extras:
                target = f"{target} {extras}" if target else str(extras)

            success = order.get("success", "")
            error = order.get("error", "")

            values = [gt, op_stage, phase, sub_phase, side,
                      cmd, unit_id, target, success, error]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value="" if v is None else v)
                _style_data_cell(cell)

            # Side colouring
            if side == "allied":
                for ci in range(1, len(values) + 1):
                    ws.cell(row=row, column=ci).fill = ROW_FILL_ALLIED
            elif side == "axis":
                for ci in range(1, len(values) + 1):
                    ws.cell(row=row, column=ci).fill = ROW_FILL_AXIS

            row += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ORDERS_HEADERS))}{row - 1}"
    return row - 2


# ---------------------------------------------------------------------------
#  Sheet 3: Turn Summary
# ---------------------------------------------------------------------------
SUMMARY_HEADERS = [
    "GT", "Weather", "Season",
    "Allied Active", "Allied SP", "Allied TOE SP", "Allied SP%",
    "Allied Water Avg%", "Allied Fuel Avg%",
    "Axis Active", "Axis SP", "Axis TOE SP", "Axis SP%",
    "Axis Water Avg%", "Axis Fuel Avg%",
    "Allied Situation", "Axis Situation",
    "Allied Orders", "Axis Orders",
    "Turn Time (s)",
]


def _aggregate_units(units, side):
    """Compute aggregate stats for one side."""
    active = 0
    sp_total = 0
    toe_total = 0
    water_pcts = []
    fuel_pcts = []

    for u in units.values():
        if u.get("side") != side:
            continue
        if u.get("status") != "active":
            continue
        active += 1
        sp = _sum_strength(u.get("current_strength"))
        toe = _sum_strength(u.get("toe_strength"))
        sp_total += sp
        toe_total += toe

        supply = u.get("supply", {})
        wc = supply.get("water_capacity", 0) or 0
        if wc > 0:
            water_pcts.append((supply.get("water", 0) or 0) / wc)
        fc = supply.get("fuel_capacity", 0) or 0
        if fc > 0:
            fuel_pcts.append((supply.get("fuel", 0) or 0) / fc)

    sp_pct = sp_total / toe_total if toe_total else 0
    avg_water = sum(water_pcts) / len(water_pcts) if water_pcts else 0
    avg_fuel = sum(fuel_pcts) / len(fuel_pcts) if fuel_pcts else 0
    return active, sp_total, toe_total, sp_pct, avg_water, avg_fuel


def _extract_situations(log_entries):
    """Build {gt: {side: situation_string}} from expert_recommendations in movement_combat phases."""
    result = {}
    for entry in log_entries:
        if entry.get("type") != "phase":
            continue
        sub = str(entry.get("sub_phase", "")).lower()
        if "movement" not in sub:
            continue
        gt = entry.get("game_turn")
        side = entry.get("side", "")
        recs = entry.get("expert_recommendations", [])
        if not recs:
            continue
        # Combine all recommendations into a summary string
        parts = []
        for r in recs:
            sit = r.get("situation", "")
            reasoning = r.get("reasoning", "")
            if sit:
                parts.append(f"{sit}: {reasoning}" if reasoning else sit)
        if parts:
            result.setdefault(gt, {})[side] = " | ".join(parts)
    return result


def _count_orders_by_turn(log_entries):
    """Build {gt: {side: count}}."""
    result = {}
    for entry in log_entries:
        if entry.get("type") != "phase":
            continue
        gt = entry.get("game_turn")
        side = entry.get("side", "")
        orders = entry.get("orders", [])
        # Don't count end_phase as a real order
        real_orders = [o for o in orders if o.get("command") != "end_phase"]
        if real_orders:
            result.setdefault(gt, {}).setdefault(side, 0)
            result[gt][side] += len(real_orders)
    return result


def _turn_times(log_entries):
    """Build {gt: elapsed_seconds} from turn_end entries."""
    result = {}
    for entry in log_entries:
        if entry.get("type") == "turn_end":
            gt = entry.get("game_turn")
            ms = entry.get("elapsed_ms", 0)
            result[gt] = round(ms / 1000, 1) if ms else ""
    return result


def write_turn_summary(wb, saves_by_turn, log_entries):
    ws = wb.create_sheet("Turn Summary")

    for ci, h in enumerate(SUMMARY_HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, HEADER_FILL_NEUTRAL, len(SUMMARY_HEADERS))

    situations = _extract_situations(log_entries)
    order_counts = _count_orders_by_turn(log_entries)
    times = _turn_times(log_entries)

    row = 2
    for gt in sorted(saves_by_turn.keys()):
        data = load_save(saves_by_turn[gt])
        if data is None:
            continue
        turn = data.get("turn", {})
        units = data.get("units", {})

        a_active, a_sp, a_toe, a_sp_pct, a_water, a_fuel = _aggregate_units(units, "allied")
        x_active, x_sp, x_toe, x_sp_pct, x_water, x_fuel = _aggregate_units(units, "axis")

        gt_sits = situations.get(gt, {})
        gt_orders = order_counts.get(gt, {})

        values = [
            gt,
            turn.get("current_weather", ""),
            turn.get("current_season", ""),
            a_active, a_sp, a_toe, a_sp_pct,
            a_water, a_fuel,
            x_active, x_sp, x_toe, x_sp_pct,
            x_water, x_fuel,
            gt_sits.get("allied", ""),
            gt_sits.get("axis", ""),
            gt_orders.get("allied", 0),
            gt_orders.get("axis", 0),
            times.get(gt, ""),
        ]

        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            _style_data_cell(cell, is_number=isinstance(v, (int, float)))

        # Percentage formatting
        for pct_col in [7, 8, 9, 13, 14, 15]:
            ws.cell(row=row, column=pct_col).number_format = PCT_FMT

        row += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{row - 1}"
    return row - 2


# ---------------------------------------------------------------------------
#  Sheet 4: Dumps
# ---------------------------------------------------------------------------
DUMPS_HEADERS = [
    "Numero", "Hex", "Side", "Water", "Fuel", "Stores", "Ammo",
]


def write_dumps(wb, saves_by_turn):
    ws = wb.create_sheet("Dumps")

    for ci, h in enumerate(DUMPS_HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, HEADER_FILL_NEUTRAL, len(DUMPS_HEADERS))

    # Use the latest save
    latest_gt = max(saves_by_turn.keys()) if saves_by_turn else None
    if latest_gt is None:
        return 0
    data = load_save(saves_by_turn[latest_gt])
    if data is None:
        return 0

    hexes = data.get("hexes", {})
    row = 2
    for hex_id in sorted(hexes.keys()):
        hex_data = hexes[hex_id]
        dumps = hex_data.get("supply_dumps", [])
        if not dumps:
            continue
        for dump in dumps:
            dump_id = dump.get("id", "")
            side = dump.get("side", "")
            values = [
                dump_id,
                hex_id,
                side,
                round(dump.get("water", 0), 2),
                round(dump.get("fuel", 0), 2),
                round(dump.get("stores", 0), 2),
                round(dump.get("ammo", 0), 2),
            ]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                _style_data_cell(cell, is_number=isinstance(v, (int, float)))

            if side == "allied":
                for ci in range(1, len(values) + 1):
                    ws.cell(row=row, column=ci).fill = ROW_FILL_ALLIED
            elif side == "axis":
                for ci in range(1, len(values) + 1):
                    ws.cell(row=row, column=ci).fill = ROW_FILL_AXIS

            row += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DUMPS_HEADERS))}{row - 1}"
    return row - 2


# ---------------------------------------------------------------------------
#  Sheet 5 (bonus): Aircraft Status  -- from latest save
# ---------------------------------------------------------------------------
AIRCRAFT_HEADERS = [
    "ID", "Side", "Type", "SGSU", "Status",
    "Current Mission", "Sorties Flown", "Ready",
]


def write_aircraft(wb, saves_by_turn):
    latest_gt = max(saves_by_turn.keys()) if saves_by_turn else None
    if latest_gt is None:
        return 0
    data = load_save(saves_by_turn[latest_gt])
    if data is None:
        return 0

    aircraft = data.get("aircraft", {})
    if not aircraft:
        return 0

    ws = wb.create_sheet("Aircraft")
    for ci, h in enumerate(AIRCRAFT_HEADERS, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, HEADER_FILL_NEUTRAL, len(AIRCRAFT_HEADERS))

    # Build SGSU lookup: aircraft_id -> sgsu_id
    sgsu_map = {}
    for sgsu_id, sgsu in data.get("sgsus", {}).items():
        for ac_id in sgsu.get("aircraft_ids", []):
            sgsu_map[ac_id] = sgsu_id

    row = 2
    for ac_id in sorted(aircraft.keys()):
        ac = aircraft[ac_id]
        values = [
            ac.get("id", ""),
            ac.get("side", ""),
            ac.get("aircraft_type_id", ac.get("type", "")),
            sgsu_map.get(ac_id, ""),
            ac.get("status", ""),
            ac.get("mission", ac.get("current_mission", "")),
            ac.get("sorties_flown", ""),
            "Y" if ac.get("status") == "ready" else "N",
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value="" if v is None else v)
            _style_data_cell(cell)

        side = ac.get("side", "")
        if side == "allied":
            for ci in range(1, len(values) + 1):
                ws.cell(row=row, column=ci).fill = ROW_FILL_ALLIED
        elif side == "axis":
            for ci in range(1, len(values) + 1):
                ws.cell(row=row, column=ci).fill = ROW_FILL_AXIS

        row += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)
    return row - 2


# ---------------------------------------------------------------------------
#  Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Export CNA game data to an Excel workbook matching reference-sheet style."
    )
    parser.add_argument(
        "--log-dir", default="logs",
        help="Directory containing game_*.jsonl log files (default: logs)",
    )
    parser.add_argument(
        "--save-dir", default="saves",
        help="Directory containing gt*.json save files (default: saves)",
    )
    parser.add_argument(
        "--output", default="output/game_report.xlsx",
        help="Output Excel file path (default: output/game_report.xlsx)",
    )
    parser.add_argument(
        "--min-gt", type=int, default=None,
        help="Minimum game turn to include (default: all)",
    )
    parser.add_argument(
        "--max-gt", type=int, default=None,
        help="Maximum game turn to include (default: all)",
    )
    parser.add_argument(
        "--game-log", default=None,
        help="Explicit path to a game_*.jsonl file (overrides auto-discovery)",
    )
    args = parser.parse_args()

    # Resolve paths relative to CWD
    log_dir = os.path.abspath(args.log_dir)
    save_dir = os.path.abspath(args.save_dir)
    output_path = os.path.abspath(args.output)

    print(f"CNA Game Report Exporter")
    print(f"  Save dir : {save_dir}")
    print(f"  Log dir  : {log_dir}")
    print(f"  Output   : {output_path}")
    print()

    # -- Discover saves --
    saves_by_turn = discover_save_files(save_dir)
    if not saves_by_turn:
        sys.exit(f"ERROR: No save files found in {save_dir}")

    # Filter by turn range if specified
    if args.min_gt is not None:
        saves_by_turn = {k: v for k, v in saves_by_turn.items() if k >= args.min_gt}
    if args.max_gt is not None:
        saves_by_turn = {k: v for k, v in saves_by_turn.items() if k <= args.max_gt}
    if not saves_by_turn:
        sys.exit(f"ERROR: No save files in the specified turn range")

    turns = sorted(saves_by_turn.keys())
    print(f"Found {len(saves_by_turn)} save files (GT {turns[0]} - GT {turns[-1]})")

    # -- Discover game log --
    log_path = args.game_log or discover_game_log(log_dir)
    log_entries = []
    if log_path:
        print(f"Using game log: {os.path.basename(log_path)}")
        log_entries = load_game_log(log_path)
        print(f"  {len(log_entries)} log entries loaded")
    else:
        print("WARNING: No game log found -- Orders and situation data will be empty")
    print()

    # -- Create workbook --
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Writing sheets...")

    n_unit = write_unit_status(wb, saves_by_turn)
    print(f"  Unit Status : {n_unit:,} rows ({len(saves_by_turn)} turns x units)")

    n_orders = write_orders(wb, log_entries)
    print(f"  Orders      : {n_orders:,} rows")

    n_summary = write_turn_summary(wb, saves_by_turn, log_entries)
    print(f"  Turn Summary: {n_summary:,} rows")

    n_dumps = write_dumps(wb, saves_by_turn)
    print(f"  Dumps       : {n_dumps:,} rows (from GT {turns[-1]})")

    n_aircraft = write_aircraft(wb, saves_by_turn)
    print(f"  Aircraft    : {n_aircraft:,} rows (from GT {turns[-1]})")

    # -- Save --
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print()
    print(f"Saved to {output_path}")
    file_size = os.path.getsize(output_path)
    if file_size > 1_000_000:
        print(f"  File size: {file_size / 1_000_000:.1f} MB")
    else:
        print(f"  File size: {file_size / 1_000:.1f} KB")


if __name__ == "__main__":
    main()
